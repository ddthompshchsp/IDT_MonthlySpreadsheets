import streamlit as st
import pandas as pd
import numpy as np
import re
import unicodedata
from io import BytesIO
from zipfile import ZipFile, ZIP_DEFLATED
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------- UI SETUP ----------------
st.set_page_config(page_title='HCHSP Department IDT Files (ZIP)', layout='wide')
st.title('Department Packets - Wide Excel ZIP (normalized)')

st.markdown("""

- Department cell turns **green** if it equals the FSW value, **red** if it differs.

**Upload:**
1) `FSW_Master` (xlsx/csv) with columns: **Month, Area, Campus, FSW, Metric, Value** (or **FSW_Value**)  
2) `metrics_map.csv` with columns: **Department, Metric** (exact labels you want as columns)

*(Optional)* a roster file (xlsx/csv) with **Area, Campus, FSW** to ensure full coverage; we also union with master so no FSW disappears.
""")

AREA_SHEETS = ['Central','West','East']
READONLY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
GREEN_FILL    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL      = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
VALID_STATUSES = ['Validated','Mismatch','Unable to Validate']

# ---------------- Helpers ----------------
def _norm(s):
    return s.replace('\u2013','-').replace('\u2014','-').strip() if isinstance(s,str) else s

def load_any(uploaded):
    if uploaded is None:
        return None
    name = getattr(uploaded, 'name', '').lower()
    if name.endswith('.csv'):
        df = pd.read_csv(uploaded)
    else:
        df = pd.read_excel(uploaded, engine='openpyxl')
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].map(_norm)
    return df

def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def normalize_metric_key(s: str) -> str:
    """Canonical key for matching metrics across files (case/spacing/punctuation safe)."""
    s = (s or "").strip().lower()
    s = s.replace("\u2013","-").replace("\u2014","-")  # en/em dash -> hyphen
    s = _strip_accents(s)
    s = re.sub(r"[^a-z0-9]+", " ", s)   # keep letters/digits as tokens
    s = re.sub(r"\s+", " ", s).strip()
    return s

def normalize_master_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize flexible headers to: Month, Area, Campus, FSW, Metric, FSW_Value."""
    lower = {str(c).strip().lower(): c for c in df.columns}

    def pick(*cands):
        for c in cands:
            if c in lower:
                return lower[c]
        return None

    col_month  = pick('month')
    col_area   = pick('area','region','cluster')
    col_campus = pick('campus','center','center/campus','site','school','location')
    col_fsw    = pick('fsw','fsw name','family service worker','family advocate','staff name')
    col_metric = pick('metric','metrics')
    col_value  = pick('value','fsw_value','fsw value')

    needed = [col_month, col_area, col_campus, col_fsw, col_metric, col_value]
    if any(c is None for c in needed):
        return df  # caller will error

    out = df.rename(columns={
        col_month:  'Month',
        col_area:   'Area',
        col_campus: 'Campus',
        col_fsw:    'FSW',
        col_metric: 'Metric',
        col_value:  'FSW_Value',
    })
    for c in ['Month','Area','Campus','FSW','Metric']:
        out[c] = out[c].astype(str).str.strip()
    # numeric coerce without forcing zeros
    out['FSW_Value'] = pd.to_numeric(out['FSW_Value'], errors='coerce')
    return out

def excel_col(n:int)->str:
    s = ''
    while n:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def add_styles_filters(ws):
    # header
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical='center')
    # freeze and autofilter
    ws.freeze_panes = 'E2'
    last_col = excel_col(ws.max_column)
    ws.auto_filter.ref = f'A1:{last_col}{ws.max_row}'

    # widths
    for i in range(1, ws.max_column+1):
        ws.column_dimensions[excel_col(i)].width = 16
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22

    # gray out ID columns (read-only visual)
    for col_letter in ['A','B','C','D']:
        for cell in ws[col_letter][1:]:
            cell.fill = READONLY_FILL

def add_match_colors(ws, metric_pairs, start_row=2, end_row=None):
    if end_row is None:
        end_row = ws.max_row
    if end_row < start_row:
        return
    # Use range-based CF to avoid per-row proliferation
    for fsw_col, dept_col in metric_pairs:
        ws.conditional_formatting.add(
            f'{dept_col}{start_row}:{dept_col}{end_row}',
            FormulaRule(
                formula=[f'AND({dept_col}{start_row}<>"", {fsw_col}{start_row}<>"", {dept_col}{start_row}<> {fsw_col}{start_row})'],
                stopIfTrue=False,
                fill=RED_FILL
            )
        )
        ws.conditional_formatting.add(
            f'{dept_col}{start_row}:{dept_col}{end_row}',
            FormulaRule(
                formula=[f'AND({dept_col}{start_row}<>"", {fsw_col}{start_row}<>"", {dept_col}{start_row}= {fsw_col}{start_row})'],
                stopIfTrue=False,
                fill=GREEN_FILL
            )
        )

def add_validations(ws, col_letter, max_row):
    if max_row < 2:
        return
    dv = DataValidation(type='list', formula1='"' + ','.join(VALID_STATUSES) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'{col_letter}2:{col_letter}{max_row}')

def ensure_roster_union(master_month_df, roster_df=None):
    """Union of FSWs from master_month_df and (optional) roster, so nobody disappears."""
    from_master = master_month_df[['Area','Campus','FSW']].drop_duplicates()
    if roster_df is not None and set(['Area','Campus','FSW']).issubset(roster_df.columns):
        from_roster = roster_df[['Area','Campus','FSW']].drop_duplicates()
        all_rows = pd.concat([from_master, from_roster], ignore_index=True).drop_duplicates()
    else:
        all_rows = from_master
    all_rows = all_rows.assign(
        Area   = all_rows['Area'].astype(str).str.strip(),
        Campus = all_rows['Campus'].astype(str).str.strip(),
        FSW    = all_rows['FSW'].astype(str).str.strip(),
    )
    return all_rows.sort_values(['Area','Campus','FSW'])

def build_wide_sheet(
    wb,
    sheet_name,
    month_value,
    roster_area_df,
    fsw_month_area_df,
    metrics_order,
    fill_missing_zero=False,  # affects FSW value columns only
):
    ws = wb.create_sheet(sheet_name)

    # Header: Month, IDs, then interleaved metric columns, then admin columns
    interleaved = []
    for m in metrics_order:
        interleaved += [m, f'Department - {m}']
    headers = ['Month','Area','Campus','FSW'] + interleaved + ['Validated','Validation_Date','Issues','Services','Referrals','Notes']
    ws.append(headers)

    # Pivot FSW values -> rows per FSW, columns per Metric (display labels)
    if not fsw_month_area_df.empty:
        pv = fsw_month_area_df.pivot_table(
            index=['Area','Campus','FSW'],
            columns='Metric',
            values='FSW_Value',
            aggfunc='first'
        )
        # Ensure all desired metric columns exist
        for m in metrics_order:
            if m not in pv.columns:
                pv[m] = np.nan
        pv = pv[metrics_order].reset_index()
    else:
        pv = pd.DataFrame(columns=['Area','Campus','FSW'] + metrics_order)

    # Base = all FSWs in this area (from union), then left-join their FSW values
    base = roster_area_df.copy() if (roster_area_df is not None and not roster_area_df.empty) else pd.DataFrame(columns=['Area','Campus','FSW'])
    if not pv.empty:
        base = base.merge(pv, on=['Area','Campus','FSW'], how='left')
    else:
        for m in metrics_order:
            base[m] = np.nan

    # Optionally zero-fill FSW columns if requested
    if fill_missing_zero and metrics_order:
        present = [m for m in metrics_order if m in base.columns]
        if present:
            base[present] = base[present].fillna(0)

    # Department entry columns (blank)
    for m in metrics_order:
        base[f'Department - {m}'] = ''

    # Trailing admin columns
    base['Validated'] = ''
    base['Validation_Date'] = ''
    base['Issues'] = ''
    base['Services'] = ''
    base['Referrals'] = ''
    base['Notes'] = ''

    # Month column
    base.insert(0, 'Month', month_value)

    # Ensure all header columns exist before ordering
    for col in ['Month','Area','Campus','FSW'] + interleaved + ['Validated','Validation_Date','Issues','Services','Referrals','Notes']:
        if col not in base.columns:
            base[col] = []

    ordered = ['Month','Area','Campus','FSW'] + interleaved + ['Validated','Validation_Date','Issues','Services','Referrals','Notes']
    base = base.loc[:, ordered]

    # Write rows (if any)
    if not base.empty:
        for r in dataframe_to_rows(base, index=False, header=False):
            ws.append(r)

    # Styles & filters
    add_styles_filters(ws)

    # Conditional formatting (Dept vs FSW)
    metric_pairs = []
    start_idx = 5  # Month..FSW = 4 columns
    for i, m in enumerate(metrics_order):
        fsw_idx  = start_idx + 2*i
        dept_idx = start_idx + 2*i + 1
        metric_pairs.append((excel_col(fsw_idx), excel_col(dept_idx)))
    if ws.max_row >= 2 and metric_pairs:
        add_match_colors(ws, metric_pairs, start_row=2, end_row=ws.max_row)

    # Validated dropdown
    validated_idx = 4 + 2*len(metrics_order) + 1
    if ws.max_row >= 2:
        add_validations(ws, excel_col(validated_idx), ws.max_row)

    return ws

def build_workbook_bytes(
    master_month_df,
    month_value,
    metrics_order,
    roster_df=None,
    fill_missing_zero=False,
):
    wb = Workbook()
    wb.remove(wb.active)

    roster_all = ensure_roster_union(master_month_df, roster_df)

    for area in AREA_SHEETS:
        roster_area = roster_all[roster_all['Area'].str.casefold() == area.lower()].copy()
        fsw_area    = master_month_df[master_month_df['Area'].astype(str).str.strip().str.casefold() == area.lower()].copy()
        build_wide_sheet(
            wb, area, month_value,
            roster_area_df=roster_area,
            fsw_month_area_df=fsw_area,
            metrics_order=metrics_order,
            fill_missing_zero=fill_missing_zero,
        )

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ---------------- INPUTS ----------------
left, right = st.columns(2)
with left:
    fsw_up = st.file_uploader('FSW_Master (xlsx/csv)', type=['xlsx','csv'])
with right:
    mm_up = st.file_uploader('metrics_map.csv (Department, Metric)', type=['csv'])

roster_up = st.file_uploader('Optional: Roster (xlsx/csv) with Area, Campus, FSW', type=['xlsx','csv'])

if fsw_up is None or mm_up is None:
    st.info('Upload FSW_Master and metrics_map to continue.')
    st.stop()

fsw_raw = load_any(fsw_up)
mmap = load_any(mm_up)
roster_df = load_any(roster_up) if roster_up is not None else None

# Normalize master columns
fsw = normalize_master_columns(fsw_raw)
need_cols = {'Month','Area','Campus','FSW','Metric','FSW_Value'}
missing = need_cols - set(fsw.columns)
if missing:
    st.error(f"FSW_Master is missing columns: {sorted(missing)}")
    st.stop()

# Clean strings
for c in ['Month','Area','Campus','FSW','Metric']:
    fsw[c] = fsw[c].astype(str).str.strip()

# Build canonical key in master
fsw['Metric_key'] = fsw['Metric'].map(normalize_metric_key)

# Validate metrics_map + keys
if not {'Department','Metric'}.issubset(mmap.columns):
    st.error('metrics_map.csv must have columns: Department, Metric')
    st.stop()
mmap['Department'] = mmap['Department'].astype(str).str.strip()
mmap['Metric']     = mmap['Metric'].astype(str).str.strip()
mmap['Metric_key'] = mmap['Metric'].map(normalize_metric_key)

# Optional warning for duplicate (Department, Metric_key)
dup_keys = mmap[mmap.duplicated(['Department','Metric_key'], keep=False)]
if not dup_keys.empty:
    st.warning("Duplicate department↔metric mappings detected (after normalization). Please check metrics_map.csv.")

# Attach Department to master via normalized key (left join keeps all FSW rows)
master = fsw.merge(mmap[['Metric_key','Department']], on='Metric_key', how='left')

# ---------------- Debug / Health ----------------
with st.expander("Debug / Data Health", expanded=False):
    st.write("Rows in master:", len(master))
    st.write("Unique Months:", sorted(master['Month'].dropna().unique().tolist()))
    st.write("Unique Campuses:", len(master['Campus'].dropna().unique()))
    st.write("Unique FSWs:", master['FSW'].nunique())
    mm_counts = mmap.groupby('Department')['Metric'].count().rename('Metric_Count').reset_index()
    st.dataframe(mm_counts, use_container_width=True)
    unmapped = master[master['Department'].isna()]['Metric'].value_counts().head(20)
    if not unmapped.empty:
        st.warning("Top unmapped metrics in master (no Department after normalization):")
        st.write(unmapped)

# ---------------- Filters ----------------
st.subheader('Filters & Options')
c1, c2, c3 = st.columns(3)

months_all = [m for m in ['Sep','Oct','Nov','Dec','Jan','Feb','Mar','Apr','May']
              if m in master['Month'].dropna().unique().tolist()]
depts_all  = sorted(mmap['Department'].dropna().unique().tolist())
camp_all   = sorted(master['Campus'].dropna().unique().tolist())

months = c1.multiselect('Months', months_all, default=months_all)
depts  = c2.multiselect('Departments', depts_all, default=depts_all)
camp   = c3.multiselect('Campuses', camp_all, default=camp_all)

fill_zero = st.checkbox('Fill missing FSW metric values with 0 (Department columns stay blank)', value=False)

mf = master.copy()
if months: mf = mf[mf['Month'].isin(months)]
if camp:   mf = mf[mf['Campus'].isin(camp)]

# ---------------- Build ZIP ----------------
if st.button('Build ZIP of Department Packets (Excel)'):
    if mf.empty:
        st.warning('No rows after filters. Check Months/Campuses.')
    else:
        memzip = BytesIO()
        with ZipFile(memzip, 'w', ZIP_DEFLATED) as zf:
            for month in sorted(mf['Month'].dropna().unique().tolist()):
                m_month = mf[mf['Month'] == month].copy()

                target_depts = depts if depts else sorted(mmap['Department'].dropna().unique())
                for dept in target_depts:
                    # Metrics for this department (display order from CSV)
                    dept_rows = mmap.loc[mmap['Department'] == dept, ['Metric','Metric_key']]
                    metrics_order = dept_rows['Metric'].tolist()           # display labels for sheet columns
                    metrics_keys  = set(dept_rows['Metric_key'].tolist())  # normalized keys for filtering

                    # Keep only FSW rows for these metrics (by normalized key)
                    month_for_metrics = m_month[m_month['Metric_key'].isin(metrics_keys)][
                        ['Month','Area','Campus','FSW','Metric','Metric_key','FSW_Value']
                    ].copy()

                    # If nothing matched, fabricate an empty frame that still lists FSWs
                    if month_for_metrics.empty:
                        base_ids = m_month[['Month','Area','Campus','FSW']].drop_duplicates().copy()
                        base_ids['Metric'] = pd.Series(dtype=str)
                        base_ids['Metric_key'] = pd.Series(dtype=str)
                        base_ids['FSW_Value'] = pd.Series(dtype=float)
                        month_for_metrics = base_ids

                    # IMPORTANT: ensure the 'Metric' column in this slice equals the display labels
                    # as defined by metrics_map for this department (so pivot aligns to metrics_order).
                    key_to_display = dict(zip(dept_rows['Metric_key'], dept_rows['Metric']))
                    month_for_metrics['Metric'] = month_for_metrics['Metric_key'].map(key_to_display)

                    xbytes = build_workbook_bytes(
                        master_month_df=month_for_metrics[['Month','Area','Campus','FSW','Metric','FSW_Value']].copy(),
                        month_value=month,
                        metrics_order=metrics_order,
                        roster_df=roster_df,
                        fill_missing_zero=fill_zero,
                    )

                    safe_dept = "".join(ch for ch in dept if ch.isalnum() or ch in (" ","-","_")).strip().replace(" ","_") or "Dept"
                    zf.writestr(f"{month}/{safe_dept}_DEPT.xlsx", xbytes)

        memzip.seek(0)
        stamp = datetime.now().strftime('%Y%m%d_%H%M')
        st.download_button(
            'Download Department Packets ZIP',
            data=memzip,
            file_name=f'DeptPackets_{stamp}.zip',
            mime='application/zip'
        )
